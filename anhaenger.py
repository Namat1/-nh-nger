import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# Titel der App
st.title("Filterung nach bestimmten Zahlen")

# Datei-Upload
uploaded_file = st.file_uploader("Lade deine Excel-Datei hoch", type=["xlsx", "xls"])

# Definierte Zahlen, nach denen gefiltert werden soll
filter_numbers = ["602", "620", "350", "520", "156"]

if uploaded_file:
    try:
        # Excel-Datei mit openpyxl laden
        workbook = load_workbook(uploaded_file, data_only=True)

        # Prüfen, ob das Blatt 'Touren' existiert
        if "Touren" in workbook.sheetnames:
            sheet = workbook["Touren"]  # Blatt 'Touren' auswählen

            # Daten in ein DataFrame umwandeln
            data = sheet.values
            columns = next(data)  # Spaltennamen extrahieren
            df = pd.DataFrame(data, columns=columns)

            # Prüfen, ob die relevante Spalte existiert
            target_column = "Unnamed: 11"  # Hier die Spalte mit den Zahlen
            if target_column in df.columns:
                # Filter auf die definierten Zahlen anwenden
                filtered_df = df[df[target_column].astype(str).isin(filter_numbers)]

                # Ergebnisse anzeigen
                st.write("Gefilterte Ergebnisse:")
                if not filtered_df.empty:
                    st.dataframe(filtered_df)

                    # Option zum Exportieren der gefilterten Daten
                    output = filtered_df.to_excel(index=False, engine="openpyxl")
                    st.download_button(
                        label="Gefilterte Daten herunterladen",
                        data=output,
                        file_name="Gefilterte_Daten.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                else:
                    st.warning("Keine Einträge gefunden, die den Filterkriterien entsprechen.")
            else:
                st.error(f"Die Spalte '{target_column}' wurde nicht in der Datei gefunden.")
        else:
            st.error("Das Blatt 'Touren' wurde in der Datei nicht gefunden.")
    except Exception as e:
        st.error(f"Fehler beim Verarbeiten der Datei: {e}")
else:
    st.info("Bitte lade eine Excel-Datei hoch, um zu starten.")
